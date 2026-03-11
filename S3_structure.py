import boto3
from collections import defaultdict
from openpyxl import Workbook
import re

MAX_ITEMS = 20
SKIP_PATTERNS = ["logging", "dev"]

s3 = boto3.client("s3")


def sanitize_sheet_name(name):
    name = re.sub(r'[:\\/?*\[\]]', '_', name)
    return name[:31]


def list_bucket_keys(bucket):
    paginator = s3.get_paginator("list_objects_v2")

    for page in paginator.paginate(Bucket=bucket):
        for obj in page.get("Contents", []):
            yield obj["Key"]


def build_tree(keys):
    tree = lambda: defaultdict(tree)
    root = tree()

    for key in keys:
        parts = key.split("/")
        node = root

        for part in parts:
            node = node[part]

    return root


def prune_tree(node):
    if len(node) > MAX_ITEMS:
        return {"EXCEEDED": True, "COUNT": len(node)}

    new_node = {}

    for k, v in node.items():
        new_node[k] = prune_tree(v)

    return new_node


def write_tree(ws, node, level=0, row=1):

    indent = "    " * level

    for name, child in node.items():

        is_folder = isinstance(child, dict)
        icon = "📁" if is_folder else "📄"

        ws.cell(row=row, column=1, value=f"{indent}{icon} {name}")
        ws.row_dimensions[row].outlineLevel = level

        row += 1

        if isinstance(child, dict):

            if child.get("EXCEEDED"):
                ws.cell(
                    row=row,
                    column=1,
                    value=f"{indent}    ⚠ exceeds {MAX_ITEMS} items"
                )
                ws.row_dimensions[row].outlineLevel = level + 1
                row += 1

            else:
                row = write_tree(ws, child, level + 1, row)

    return row


def process_bucket(bucket, wb):

    print(f"Scanning bucket: {bucket}")

    keys = list(list_bucket_keys(bucket))

    tree = build_tree(keys)

    pruned_tree = prune_tree(tree)

    ws = wb.create_sheet(title=sanitize_sheet_name(bucket))

    ws.sheet_properties.outlinePr.summaryBelow = True
    ws.sheet_view.showOutlineSymbols = True

    ws.column_dimensions["A"].width = 120

    ws.cell(row=1, column=1, value="S3 Bucket Structure")

    write_tree(ws, pruned_tree, level=0, row=2)


def main():

    wb = Workbook()
    wb.remove(wb.active)

    all_buckets = s3.list_buckets()["Buckets"]

    for b in all_buckets:

        bucket = b["Name"]

        if any(p in bucket.lower() for p in SKIP_PATTERNS):
            print(f"Skipping bucket: {bucket}")
            continue

        process_bucket(bucket, wb)

    wb.save("s3_bucket_structure.xlsx")

    print("Excel file created: s3_bucket_structure.xlsx")


if __name__ == "__main__":
    main()
